import pandas as pd
import os
from openpyxl import load_workbook

def get_column(df, possible_names, default_value=None):
    """
    Возвращает значение столбца, если он существует, или заполняет столбец значением по умолчанию.
    """
    for name in possible_names:
        if name in df.columns:
            return df[name]
    return pd.Series([default_value] * len(df))

def get_price(df):
    return get_column(df, ['оптовая цена, руб.', 'ценв rub', 'цена, руб', 'розница руб.', 'цена партнера', 'price'], None)

def get_vendor(df):
    return get_column(df, ['марка', 'производитель', 'бренд'], 'Не указан')

def get_sklad(df):
    return get_column(df, ['город', 'склад'], 'Москва')

def process_file(file_path):
    try:
        # Загружаем данные
        df = pd.read_excel(file_path)
        df.columns = [col.strip().lower() for col in df.columns]

        # Получаем название файла для столбца "Поставщик"
        manufacturer = os.path.splitext(os.path.basename(file_path))[0]

        # Заполняем необходимые столбцы
        df['Стоимость'] = get_price(df)
        df['Поставщик'] = manufacturer
        df['Вендор'] = get_vendor(df)
        df['Артикул'] = get_column(df, ['артикул'], 'Не указан')
        df['Наименование'] = get_column(df, ['наименование'], 'Не указано')
        df['Ресурс печати'] = get_column(df, ['макс кол-во отпечатков'], 0)
        df['Количество на складе'] = get_column(df, ['кол-во', 'наличие'], 0)
        df['Склад'] = get_sklad(df)

        # Выбираем только нужные столбцы
        result_df = df[['Поставщик', 'Вендор', 'Артикул', 'Наименование', 'Стоимость',
                        'Ресурс печати', 'Количество на складе', 'Склад']]
        return result_df
    except Exception as e:
        print(f"Ошибка при обработке файла {file_path}: {e}")
        return pd.DataFrame()

# Основной процесс обработки всех файлов
file_path = 'vendorData/Прайс ZipZip Оригинад.xlsx'
processed_df = process_file(file_path)
output_file = './обработанные_данные.xlsx'
sheet_name = 'Sheet1'

book = load_workbook(output_file)

if sheet_name in book.sheetnames:
    existing_df = pd.read_excel(output_file, sheet_name=sheet_name)
    updated_df = pd.concat([existing_df, processed_df], ignore_index=True)
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
else:
    print(f"Лист '{sheet_name}' не найден в файле.")



print(f"Обработка завершена. Данные сохранены в '{output_file}'.")








